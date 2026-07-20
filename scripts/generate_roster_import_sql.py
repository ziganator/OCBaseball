#!/usr/bin/env python3
import argparse
import re
import unicodedata
from collections import Counter
from datetime import date
from pathlib import Path

from openpyxl import load_workbook


NOTE_PATTERNS = [
    "Video Forecast",
    "No new player Notes",
    "No new player Note",
    "New Player Note",
    "Player Note",
]
STATUS_RE = re.compile(r"(?:DTD|IL\d+|NA|O|SUSP|PUP)$")
TEAM_POSITION_RE = re.compile(r"([A-Z]{2,4})\s+-\s+([A-Za-z0-9,/ ]+)$")
POSITION_MAP = {
    "UTIL": "DH",
}
MLB_TEAM_NAMES = {
    "ATH": "Athletics",
    "ATL": "Atlanta Braves",
    "AZ": "Arizona Diamondbacks",
    "BAL": "Baltimore Orioles",
    "BOS": "Boston Red Sox",
    "CHC": "Chicago Cubs",
    "CIN": "Cincinnati Reds",
    "CLE": "Cleveland Guardians",
    "COL": "Colorado Rockies",
    "CWS": "Chicago White Sox",
    "DET": "Detroit Tigers",
    "HOU": "Houston Astros",
    "KC": "Kansas City Royals",
    "LAA": "Los Angeles Angels",
    "LAD": "Los Angeles Dodgers",
    "MIA": "Miami Marlins",
    "MIL": "Milwaukee Brewers",
    "MIN": "Minnesota Twins",
    "NYM": "New York Mets",
    "NYY": "New York Yankees",
    "PHI": "Philadelphia Phillies",
    "PIT": "Pittsburgh Pirates",
    "SD": "San Diego Padres",
    "SEA": "Seattle Mariners",
    "SF": "San Francisco Giants",
    "STL": "St. Louis Cardinals",
    "TB": "Tampa Bay Rays",
    "TEX": "Texas Rangers",
    "TOR": "Toronto Blue Jays",
    "WSH": "Washington Nationals",
}


def sql_string(value):
    if value is None:
        return "NULL"
    return "'" + str(value).replace("'", "''") + "'"


def sql_array(values):
    return "ARRAY[" + ", ".join(sql_string(value) for value in values) + "]::text[]"


def clean_text(value):
    text = str(value or "")
    text = "".join(ch for ch in text if not 0xE000 <= ord(ch) <= 0xF8FF)
    return text.replace("\r", "\n")


def normalize_name(name):
    text = unicodedata.normalize("NFKD", name)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"[^a-zA-Z0-9]+", " ", text).strip().lower()


def parse_player(raw_value):
    lines = [line.strip() for line in clean_text(raw_value).split("\n") if line.strip()]
    first_line = lines[0] if lines else ""
    team_position = TEAM_POSITION_RE.search(first_line)
    if not team_position:
        raise ValueError(f"Could not find MLB team/positions in: {raw_value!r}")

    mlb_abbreviation = team_position.group(1)
    positions = [
        POSITION_MAP.get(position.strip().upper(), position.strip().upper())
        for position in re.split(r"[,/]", team_position.group(2))
        if position.strip()
    ]

    name = first_line[: team_position.start()].strip()
    for note_pattern in NOTE_PATTERNS:
        name = name.replace(note_pattern, "")

    status = ""
    status_match = STATUS_RE.search(name)
    if status_match:
        status = status_match.group(0)
        name = name[: status_match.start()].strip()

    name = re.sub(r"\s+", " ", name).strip()
    if not name:
        raise ValueError(f"Could not find player name in: {raw_value!r}")

    return {
        "display_name": name,
        "normalized_name": normalize_name(name),
        "mlb_abbreviation": mlb_abbreviation,
        "primary_position": positions[0] if positions else None,
        "eligible_positions": positions,
        "status": status,
    }


def load_roster_rows(input_path):
    workbook = load_workbook(input_path, data_only=True)
    worksheet = workbook.active
    rows = []
    errors = []

    for row_number in range(2, worksheet.max_row + 1):
        raw_player = worksheet.cell(row_number, 1).value
        team_nickname = worksheet.cell(row_number, 2).value
        if raw_player is None and team_nickname is None:
            continue
        try:
            player = parse_player(raw_player)
            player["source_row"] = row_number
            player["team_nickname"] = str(team_nickname or "").strip().upper()
            rows.append(player)
        except Exception as exc:
            errors.append((row_number, str(exc)))

    if errors:
        formatted = "\n".join(f"row {row}: {message}" for row, message in errors[:20])
        raise ValueError(f"Could not parse {len(errors)} row(s):\n{formatted}")

    duplicate_names = [name for name, count in Counter(row["normalized_name"] for row in rows).items() if count > 1]
    if duplicate_names:
        raise ValueError("Duplicate player names in source: " + ", ".join(duplicate_names))

    return rows


def values_block(rows, columns):
    lines = []
    for row in rows:
        lines.append("(" + ", ".join(columns(row)) + ")")
    return ",\n  ".join(lines)


def generate_sql(input_path, output_path, season, league, acquired_on):
    rows = load_roster_rows(input_path)
    source_name = Path(input_path).name
    league_label = league.strip()
    team_counts = Counter(row["team_nickname"] for row in rows)
    status_counts = Counter(row["status"] or "ACTIVE" for row in rows)
    mlb_abbreviations = sorted({row["mlb_abbreviation"] for row in rows})
    team_rows = [
        {"abbreviation": abbreviation, "name": MLB_TEAM_NAMES.get(abbreviation, abbreviation)}
        for abbreviation in mlb_abbreviations
    ]

    player_rows = sorted(rows, key=lambda item: item["normalized_name"])
    roster_rows = sorted(rows, key=lambda item: (item["team_nickname"], item["normalized_name"]))

    team_values = values_block(
        team_rows,
        lambda row: [sql_string(row["abbreviation"]), sql_string(row["name"]), "TRUE"],
    )
    player_values = values_block(
        player_rows,
        lambda row: [
            sql_string(row["display_name"]),
            sql_string(row["normalized_name"]),
            sql_string(row["mlb_abbreviation"]),
            sql_string(row["primary_position"]),
            sql_array(row["eligible_positions"]),
        ],
    )
    roster_values = values_block(
        roster_rows,
        lambda row: [
            sql_string(row["team_nickname"]),
            sql_string(row["normalized_name"]),
            str(row["source_row"]),
        ],
    )

    header_lines = [
        f"-- Generated from {source_name}.",
        f"-- Season {season} {league_label} roster snapshot.",
        f"-- Player rows: {len(rows)}.",
        "-- Team counts: "
        + ", ".join(f"{team}: {count}" for team, count in sorted(team_counts.items())),
        "-- Yahoo status counts: "
        + ", ".join(f"{status}: {count}" for status, count in sorted(status_counts.items())),
        "-- Run after supabase/seed_teams.sql and supabase/roster_management.sql.",
    ]

    sql = f"""{"\n".join(header_lines)}

BEGIN;

INSERT INTO mlb_teams (abbreviation, name, active)
VALUES
  {team_values}
ON CONFLICT (abbreviation) DO UPDATE SET
  name = EXCLUDED.name,
  active = TRUE,
  updated_at = NOW();

WITH player_seed (
  display_name,
  normalized_name,
  mlb_abbreviation,
  primary_position,
  eligible_positions
) AS (
  VALUES
  {player_values}
)
INSERT INTO mlb_players (
  display_name,
  normalized_name,
  primary_position,
  eligible_positions,
  mlb_team_id,
  active
)
SELECT
  player_seed.display_name,
  player_seed.normalized_name,
  player_seed.primary_position,
  player_seed.eligible_positions,
  mlb_teams.id,
  TRUE
FROM player_seed
JOIN mlb_teams ON mlb_teams.abbreviation = player_seed.mlb_abbreviation
ON CONFLICT (normalized_name) DO UPDATE SET
  display_name = EXCLUDED.display_name,
  primary_position = EXCLUDED.primary_position,
  eligible_positions = EXCLUDED.eligible_positions,
  mlb_team_id = EXCLUDED.mlb_team_id,
  active = TRUE,
  updated_at = NOW();

WITH current_season AS (
  SELECT id FROM seasons WHERE season_number = {season}
)
DELETE FROM roster_memberships
USING current_season
WHERE roster_memberships.season_id = current_season.id
  AND roster_memberships.league_code = {sql_string(league_label)}
  AND roster_memberships.acquisition_type = 'admin'
  AND roster_memberships.acquired_on = DATE {sql_string(acquired_on)}
  AND roster_memberships.released_on IS NULL;

WITH current_season AS (
  SELECT id FROM seasons WHERE season_number = {season}
)
UPDATE roster_memberships
SET
  released_on = DATE {sql_string(acquired_on)},
  release_type = 'admin',
  updated_at = NOW()
FROM current_season
WHERE roster_memberships.season_id = current_season.id
  AND roster_memberships.league_code = {sql_string(league_label)}
  AND roster_memberships.released_on IS NULL;

WITH current_season AS (
  SELECT id FROM seasons WHERE season_number = {season}
),
roster_seed (team_nickname, normalized_name, source_row) AS (
  VALUES
  {roster_values}
)
INSERT INTO roster_memberships (
  season_id,
  team_id,
  league_code,
  player_id,
  acquired_on,
  acquisition_type
)
SELECT
  current_season.id,
  teams.id,
  {sql_string(league_label)},
  mlb_players.id,
  DATE {sql_string(acquired_on)},
  'admin'
FROM roster_seed
CROSS JOIN current_season
JOIN teams ON UPPER(teams.nickname) = roster_seed.team_nickname
JOIN season_team_assignments sta
  ON sta.season_id = current_season.id
  AND sta.team_id = teams.id
  AND LOWER(sta.league_code) = LOWER({sql_string(league_label)})
JOIN mlb_players ON mlb_players.normalized_name = roster_seed.normalized_name;

COMMIT;

SELECT
  teams.name AS team_name,
  COUNT(roster_memberships.id) AS active_players
FROM roster_memberships
JOIN seasons ON seasons.id = roster_memberships.season_id
JOIN teams ON teams.id = roster_memberships.team_id
WHERE seasons.season_number = {season}
  AND roster_memberships.league_code = {sql_string(league_label)}
  AND roster_memberships.released_on IS NULL
GROUP BY teams.name
ORDER BY teams.name;
"""

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(sql, encoding="utf-8")
    return rows


def main():
    parser = argparse.ArgumentParser(description="Generate a Supabase roster snapshot import from a Yahoo roster workbook.")
    parser.add_argument("input", help="Path to the Yahoo roster workbook.")
    parser.add_argument("output", help="Path to write the generated SQL file.")
    parser.add_argument("--season", type=int, default=32)
    parser.add_argument("--league", default="Keystone")
    parser.add_argument("--acquired-on", default=date.today().isoformat())
    args = parser.parse_args()

    rows = generate_sql(args.input, args.output, args.season, args.league, args.acquired_on)
    counts = Counter(row["team_nickname"] for row in rows)
    print(f"Wrote {args.output}")
    print(f"Imported rows: {len(rows)}")
    for team, count in sorted(counts.items()):
        print(f"{team}: {count}")


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
import argparse
import json
from collections import defaultdict
from datetime import date, timedelta
from pathlib import Path

from openpyxl import load_workbook


DAY_KEYS = ["M", "Tu", "W", "Th", "F", "St", "Su"]
DAY_COLS = list(range(5, 12))
FIRST_TEAM_ROWS = list(range(3, 18))
SECOND_TEAM_ROWS = list(range(25, 40))
HITTER_ROW_LIMIT = 9


def sql_string(value):
    if value is None:
        return "NULL"
    return "'" + str(value).replace("'", "''") + "'"


def sql_number(value):
    if value is None or value == "":
        return "0"
    number = float(value)
    if number.is_integer():
        return str(int(number))
    return str(number)


def json_sql(value):
    return f"{sql_string(json.dumps(value, ensure_ascii=False, separators=(',', ':')))}::jsonb"


def clean_team(value):
    return str(value or "").strip().upper()


def clean_position(value):
    text = str(value or "").strip().upper()
    return "1B" if text == "1B" else text


def points(value):
    if value is None or value == "":
        return 0
    return int(float(value))


def fill_key(cell):
    fill = cell.fill
    if fill.fill_type != "solid":
        return "primary"
    color = fill.fgColor
    if color.type == "rgb" and color.rgb:
        return color.rgb
    if color.type == "theme":
        return f"theme:{color.theme}:{color.tint}"
    return "primary"


def player_for_day(row_name, styled_row_cells, day_index):
    names = [name.strip() for name in str(row_name or "").split("/") if name.strip()]
    if len(names) <= 1:
        return names[0] if names else ""

    color_order = []
    for cell in styled_row_cells:
        key = fill_key(cell)
        if key == "primary":
            continue
        if key not in color_order:
            color_order.append(key)

    key = fill_key(styled_row_cells[day_index])
    if key == "primary":
        return names[0]
    try:
        color_index = color_order.index(key)
    except ValueError:
        return names[0]
    return names[min(color_index + 1, len(names) - 1)]


def parse_scoreboard(ws):
    matchups = []
    for row in range(5, 21):
        away = clean_team(ws.cell(row, 4).value)
        home = clean_team(ws.cell(row, 8).value)
        if not away or not home:
            continue
        matchups.append({
            "matchup_key": ws.title if ws.title != "Scoreboard" else "",
            "league_code": "keystone" if row <= 12 else "diamond",
            "away_team_name": away,
            "home_team_name": home,
            "away_score": points(ws.cell(row, 3).value),
            "home_score": points(ws.cell(row, 9).value),
            "away_psr": points(ws.cell(row, 2).value),
            "home_psr": points(ws.cell(row, 10).value),
            "away_lead": points(ws.cell(row, 5).value),
            "home_lead": points(ws.cell(row, 7).value),
        })
    return matchups


def parse_team_block(values_ws, styles_ws, rows, team_name, home_away, matchup_key, dates, source):
    player_rows = []
    team_daily = defaultdict(int)
    offense_daily = defaultdict(int)
    pitching_daily = defaultdict(int)

    for offset, row in enumerate(rows):
        position = clean_position(values_ws.cell(row, 2).value)
        row_player = str(values_ws.cell(row, 3).value or "").strip()
        if not position or not row_player:
            continue

        group = "hitter" if offset < HITTER_ROW_LIMIT else "pitcher"
        styled_cells = [styles_ws.cell(row, col) for col in DAY_COLS]

        for day_index, col in enumerate(DAY_COLS):
            day_key = DAY_KEYS[day_index]
            score = points(values_ws.cell(row, col).value)
            stat_date = dates[day_index].isoformat()
            credited_player = player_for_day(row_player, styled_cells, day_index)
            team_daily[stat_date] += score
            if group == "hitter":
                offense_daily[stat_date] += score
            else:
                pitching_daily[stat_date] += score
            player_rows.append({
                "matchup_key": matchup_key,
                "team_name": team_name,
                "player_name": credited_player,
                "roster_slot": position,
                "stat_date": stat_date,
                "sheet_points": score,
                "calculated_points": score,
                "scoring_breakdown": [],
                "raw_stats": {
                    "source": source,
                    "rowPlayer": row_player,
                    "group": group,
                    "day": day_key,
                },
            })

    team_rows = []
    for day_index, stat_day in enumerate(dates):
        stat_date = stat_day.isoformat()
        team_rows.append({
            "matchup_key": matchup_key,
            "team_name": team_name,
            "home_away": home_away,
            "stat_date": stat_date,
            "sheet_points": team_daily[stat_date],
            "calculated_points": team_daily[stat_date],
            "offense_points": offense_daily[stat_date],
            "pitching_points": pitching_daily[stat_date],
            "payload": {
                "source": source,
                "day": DAY_KEYS[day_index],
            },
        })

    return team_rows, player_rows


def generate_sql(input_path, output_path, season, game, week, start_date):
    source = Path(input_path).name
    dates = [start_date + timedelta(days=index) for index in range(7)]
    values_wb = load_workbook(input_path, data_only=True)
    styles_wb = load_workbook(input_path, data_only=False)

    scoreboard = parse_scoreboard(values_wb["Scoreboard"])
    sheet_matchups = [name for name in values_wb.sheetnames if "-" in name and name not in {"DATA"}]
    if len(scoreboard) != len(sheet_matchups):
        raise ValueError(f"Expected {len(scoreboard)} matchup sheets, found {len(sheet_matchups)}")

    matchups = []
    team_rows = []
    player_rows = []

    for index, sheet_name in enumerate(sheet_matchups):
        values_ws = values_wb[sheet_name]
        styles_ws = styles_wb[sheet_name]
        matchup = dict(scoreboard[index])
        matchup["matchup_key"] = sheet_name
        matchups.append(matchup)

        away_team_rows, away_player_rows = parse_team_block(
            values_ws, styles_ws, FIRST_TEAM_ROWS, matchup["away_team_name"], "A", sheet_name, dates, source
        )
        home_team_rows, home_player_rows = parse_team_block(
            values_ws, styles_ws, SECOND_TEAM_ROWS, matchup["home_team_name"], "H", sheet_name, dates, source
        )
        team_rows.extend(away_team_rows)
        team_rows.extend(home_team_rows)
        player_rows.extend(away_player_rows)
        player_rows.extend(home_player_rows)

    seen = set()
    duplicates = []
    for row in player_rows:
        key = (
            season,
            game,
            row["matchup_key"],
            row["team_name"],
            row["player_name"],
            row["roster_slot"],
            row["stat_date"],
        )
        if key in seen:
            duplicates.append(key)
        seen.add(key)
    if duplicates:
        raise ValueError(f"Duplicate player result keys detected: {duplicates[:10]}")

    lines = [
        f"-- Week {week} / Game {game} sheet-derived score load.",
        "-- Run after supabase/game_score_storage.sql.",
        "",
        "BEGIN;",
        "",
        f"DELETE FROM game_player_daily_score_results WHERE season_number = {season} AND game_number = {game} AND week_number = {week};",
        f"DELETE FROM game_team_daily_score_results WHERE season_number = {season} AND game_number = {game} AND week_number = {week};",
        f"DELETE FROM game_matchup_score_results WHERE season_number = {season} AND game_number = {game} AND week_number = {week};",
        f"DELETE FROM game_score_runs WHERE season_number = {season} AND game_number = {game} AND week_number = {week};",
        "",
        "INSERT INTO game_score_runs (season_number, game_number, week_number, source, status, completed_at, metadata)",
        (
            f"VALUES ({season}, {game}, {week}, 'sheet_import', 'complete', NOW(), "
            f"{json_sql({'sourceFile': source, 'startDate': dates[0].isoformat(), 'endDate': dates[-1].isoformat()})});"
        ),
        "",
        "INSERT INTO game_matchup_score_results (season_number, game_number, week_number, matchup_key, league_code, away_team_name, home_team_name, away_score, home_score, away_psr, home_psr, away_lead, home_lead, payload) VALUES",
    ]
    matchup_values = []
    for row in matchups:
        matchup_values.append(
            "("
            f"{season}, {game}, {week}, {sql_string(row['matchup_key'])}, {sql_string(row['league_code'])}, "
            f"{sql_string(row['away_team_name'])}, {sql_string(row['home_team_name'])}, "
            f"{sql_number(row['away_score'])}, {sql_number(row['home_score'])}, {sql_number(row['away_psr'])}, {sql_number(row['home_psr'])}, "
            f"{sql_number(row['away_lead'])}, {sql_number(row['home_lead'])}, {json_sql({'source': source})}"
            ")"
        )
    lines.append(",\n".join(matchup_values) + ";")

    lines.extend([
        "",
        "INSERT INTO game_team_daily_score_results (season_number, game_number, week_number, matchup_key, team_name, home_away, stat_date, sheet_points, calculated_points, offense_points, pitching_points, payload) VALUES",
    ])
    team_values = []
    for row in team_rows:
        team_values.append(
            "("
            f"{season}, {game}, {week}, {sql_string(row['matchup_key'])}, {sql_string(row['team_name'])}, {sql_string(row['home_away'])}, "
            f"{sql_string(row['stat_date'])}, {sql_number(row['sheet_points'])}, {sql_number(row['calculated_points'])}, "
            f"{sql_number(row['offense_points'])}, {sql_number(row['pitching_points'])}, {json_sql(row['payload'])}"
            ")"
        )
    lines.append(",\n".join(team_values) + ";")

    lines.extend([
        "",
        "INSERT INTO game_player_daily_score_results (season_number, game_number, week_number, matchup_key, team_name, player_name, roster_slot, stat_date, sheet_points, calculated_points, scoring_breakdown, raw_stats) VALUES",
    ])
    player_values = []
    for row in player_rows:
        player_values.append(
            "("
            f"{season}, {game}, {week}, {sql_string(row['matchup_key'])}, {sql_string(row['team_name'])}, "
            f"{sql_string(row['player_name'])}, {sql_string(row['roster_slot'])}, {sql_string(row['stat_date'])}, "
            f"{sql_number(row['sheet_points'])}, {sql_number(row['calculated_points'])}, "
            f"{json_sql(row['scoring_breakdown'])}, {json_sql(row['raw_stats'])}"
            ")"
        )
    lines.append(",\n".join(player_values) + ";")
    lines.extend(["", "COMMIT;", ""])

    Path(output_path).write_text("\n".join(lines), encoding="utf-8")
    return len(matchups), len(team_rows), len(player_rows)


def main():
    parser = argparse.ArgumentParser(description="Generate Owners Club game result SQL from a weekly workbook.")
    parser.add_argument("input")
    parser.add_argument("output")
    parser.add_argument("--season", type=int, default=32)
    parser.add_argument("--game", type=int, required=True)
    parser.add_argument("--week", type=int, required=True)
    parser.add_argument("--week-one-start", default="2026-03-30")
    args = parser.parse_args()

    week_one_start = date.fromisoformat(args.week_one_start)
    start_date = week_one_start + timedelta(days=(args.week - 1) * 7)
    counts = generate_sql(args.input, args.output, args.season, args.game, args.week, start_date)
    print(f"Wrote {args.output}: {counts[0]} matchups, {counts[1]} team-day rows, {counts[2]} player-day rows")


if __name__ == "__main__":
    main()
